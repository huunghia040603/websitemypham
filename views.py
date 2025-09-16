from rest_framework import viewsets, mixins, status, filters, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import *
from .serializers import *
from .pagination import *
from rest_framework.permissions import IsAuthenticated, AllowAny
from .filters import *
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.generics import GenericAPIView
from rest_framework import permissions
from rest_framework.decorators import action,api_view,permission_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import random
import logging
import string
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, get_connection
from django.conf import settings
from django.contrib.auth import get_user_model
from rest_framework_simplejwt.tokens import RefreshToken
from django.shortcuts import render
from rest_framework.authentication import TokenAuthentication
User = get_user_model()


class CurrentUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)



class UserInfoAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        # 'request.user' là đối tượng người dùng đã được xác thực
        user = request.user
        return Response({
            'user_id': user.id,
            'username': user.username,
            # Thêm các thông tin cần thiết khác
        })

logger = logging.getLogger(__name__)

# --- User ViewSets ---
class UserViewSet(viewsets.GenericViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    @action(detail=False, methods=['post'], url_path='forgot-password', serializer_class=ForgotPasswordSerializer)
    def forgot_password(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            logger.info(f"Yêu cầu đặt lại mật khẩu cho email không tồn tại: {email}")
            return Response(
                {"detail": "Nếu email của bạn tồn tại trong hệ thống, chúng tôi đã gửi một mã đặt lại mật khẩu."},
                status=status.HTTP_200_OK
            )

        chars = string.ascii_letters + string.digits
        new_password = ''.join(random.choice(chars) for _ in range(8))
        user.set_password(new_password)
        user.save()

        context = {
            'new_password': new_password,
        }

        try:
            html_content = render_to_string('emails/forgot_password_email.html', context)
            subject = 'Mã Đặt Lại Mật Khẩu Của Bạn'
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [email]

            with get_connection() as connection:
                msg = EmailMessage(
                    subject,
                    html_content,
                    from_email,
                    recipient_list,
                    connection=connection,
                )
                msg.content_subtype = "html"
                msg.send()

            logger.info(f"Email đặt lại mật khẩu đã được gửi thành công đến {email}")

        except Exception as e:
            # Ghi lại toàn bộ lỗi, bao gồm cả traceback
            logger.exception(f"Lỗi khi gửi email đặt lại mật khẩu đến {email}: {e}")
            return Response(
                {"detail": "Đã có lỗi xảy ra khi gửi email, vui lòng thử lại sau."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {"detail": "Một mã đặt lại mật khẩu đã được gửi đến email của bạn."},
            status=status.HTTP_200_OK
        )


class CollaboratorViewSet(viewsets.ModelViewSet):
    queryset = Collaborator.objects.all()
    serializer_class = CollaboratorSerializer

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer


# --- Product-related ViewSets ---
class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']

    def get_queryset(self):
        """
        Tự động cập nhật trạng thái của tag trước khi trả về queryset.
        """
        now = timezone.now()
        for tag in self.queryset:
            if tag.end_date < now and tag.status != 'expired':
                tag.status = 'expired'
                tag.save(update_fields=['status'])

        return self.queryset




class GiftViewSet(viewsets.ModelViewSet):
    queryset = Gift.objects.all()
    serializer_class = GiftSerializer


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.filter(is_visible=True).select_related('brand', 'category').prefetch_related('tags', 'gifts')
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProductFilter
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    # Thêm các trường tìm kiếm theo tên của brand, category và tags
    search_fields = ['name', 'brand__name', 'category__name', 'tags__name']

    # Thêm các trường sắp xếp
    ordering_fields = ['name', 'id', 'original_price', 'sold_quantity']


class AdminProductViewSet(viewsets.ModelViewSet):
    """
    ViewSet cho admin - có thể xem và quản lý tất cả sản phẩm (bao gồm cả ẩn)
    """
    queryset = Product.objects.select_related('brand', 'category').prefetch_related('tags', 'gifts')
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProductFilter
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    # Thêm các trường tìm kiếm theo tên của brand, category và tags
    search_fields = ['name', 'brand__name', 'category__name', 'tags__name']

    # Thêm các trường sắp xếp
    ordering_fields = ['name', 'id', 'original_price', 'sold_quantity', 'is_visible']

    @action(detail=False, methods=['patch'])
    def add_tag(self, request):
        tag_id = request.data.get('tag_id')
        product_ids = request.data.get('product_ids')

        if not tag_id or not product_ids:
            return Response(
                {"detail": "Vui lòng cung cấp 'tag_id' và 'product_ids'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            tag = Tag.objects.get(pk=tag_id)
        except Tag.DoesNotExist:
            return Response(
                {"detail": "Tag không tồn tại."},
                status=status.HTTP_404_NOT_FOUND
            )

        products = self.get_queryset().filter(pk__in=product_ids)
        if not products.exists():
            return Response(
                {"detail": "Không tìm thấy sản phẩm nào để cập nhật."},
                status=status.HTTP_404_NOT_FOUND
            )

        for product in products:
            product.tags.add(tag)

        return Response(
            {"detail": f"Đã thêm tag '{tag.name}' vào {len(products)} sản phẩm thành công."},
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['post'], url_path='import-xlsx')
    def import_xlsx(self, request):
        """Import sản phẩm từ file .xlsx (multipart/form-data, field name: file).

        - Tạo/cập nhật Brand, Category theo tên
        - Tạo Tag nếu chưa có (code sinh từ tên)
        - Cập nhật/ghi đè Product theo ID (nếu có cột id) hoặc theo tên
        - Tự tính discounted_price/discount_rate nếu thiếu 1 trong 2
        """
        from openpyxl import load_workbook
        from django.db import transaction
        from decimal import Decimal
        import re

        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Vui lòng upload file .xlsx với field name = file'}, status=status.HTTP_400_BAD_REQUEST)

        def normalize(text: str) -> str:
            if not text:
                return ''
            t = str(text).strip().lower()
            t = t.replace('\u00a0', ' ')
            return re.sub(r'\s+', ' ', t)

        # Hỗ trợ nhiều tiêu đề cột (mở rộng alias để khớp với nhiều file mẫu)
        header_aliases = {
            'id': {'id', 'mã', 'ma', 'product_id', 'product id'},
            'name': {'tên sản phẩm', 'ten san pham', 'name', 'tên', 'ten', 'product name', 'item name'},
            'brand': {'thương hiệu', 'thuong hieu', 'brand', 'hãng', 'hang', 'brand name', 'brand_name'},
            'brand_id': {'brand_id', 'brand id', 'id brand', 'mã hãng', 'ma hang', 'mã brand', 'ma brand'},
            'category': {'danh mục', 'danh muc', 'category', 'category name', 'category_name'},
            'category_id': {'category_id', 'category id', 'id category', 'mã danh mục', 'ma danh muc'},
            'import_price': {'giá nhập', 'gia nhap', 'import price', 'import_price', 'import (vnd)', 'import(vnd)', 'import'},
            'original_price': {'giá gốc', 'gia goc', 'original price', 'original_price', 'list price', 'giá niêm yết', 'gia niem yet'},
            'discounted_price': {'giá bán', 'gia ban', 'giá sau giảm', 'gia sau giam', 'discounted price', 'discounted_price', 'sale price', 'sale', 'price', 'price (vnd)', 'giá (vnd)'},
            'discount_rate': {'giảm %', 'giam %', 'discount %', 'discount_rate', 'discount rate', '% giảm', '% giam'},
            'stock_quantity': {'tồn kho', 'ton kho', 'stock', 'stock_quantity', 'số lượng', 'so luong', 'quantity', 'qty', 'stock qty'},
            'sold_quantity': {'đã bán', 'da ban', 'sold', 'sold_quantity', 'sold qty'},
            'rating': {'đánh giá', 'danh gia', 'rating', 'ratings'},
            'status': {'trạng thái', 'trang thai', 'status', 'tình trạng', 'tinh trang'},
            'image': {'ảnh', 'anh', 'image', 'image url', 'image_url', 'hình ảnh', 'hinh anh', 'thumbnail', 'image link', 'image href'},
            'description': {'mô tả', 'mo ta', 'description', 'desc', 'mô tả (html)', 'description (html)'},
            'ingredients': {'thành phần', 'thanh phan', 'ingredients', 'ingredient'},
            'tags': {'tags', 'nhãn', 'nhan', 'tag', 'labels', 'label'},
        }

        def match_key(h: str) -> str:
            h_norm = normalize(h)
            for key, aliases in header_aliases.items():
                if h_norm in aliases:
                    return key
            return h_norm  # fallback

        def to_int(val, default=0):
            try:
                if val is None or val == '':
                    return default
                return int(Decimal(str(val)))
            except Exception:
                return default

        def to_float(val, default=0.0):
            try:
                if val is None or val == '':
                    return default
                return float(Decimal(str(val)))
            except Exception:
                return default

        def slugify_code(name: str) -> str:
            base = normalize(name)
            base = re.sub(r'[^a-z0-9]+', '-', base).strip('-')
            return base[:50] or 'tag'

        try:
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'detail': f'Lỗi đọc file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Đọc header
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'File trống'}, status=status.HTTP_400_BAD_REQUEST)
        headers = rows[0]
        key_indexes = {}
        for idx, h in enumerate(headers):
            if h is None:
                continue
            key = match_key(h)
            key_indexes[key] = idx

        # Bắt buộc có cột name, brand_id, category_id để liên kết đúng brand/category
        required = ['name', 'brand_id', 'category_id']
        for r in required:
            if r not in key_indexes:
                return Response({'detail': f'Thiếu cột bắt buộc: {r}', 'detected_headers': [str(h or '') for h in headers]}, status=status.HTTP_400_BAD_REQUEST)

        created, updated, errors = 0, 0, []

        # Chuẩn hóa trạng thái từ văn bản tự do về các lựa chọn hợp lệ trong model
        allowed_status = {
            'new','30','50','70','80','85','90','95','test',
            'newmh','newm','newrt','newmn','newx','newspx','chiet'
        }

        def map_status(raw):
            s = normalize(str(raw or 'new'))
            if s in allowed_status:
                return s
            mapping = {
                'moi': 'new', 'mới': 'new', 'new nguyen': 'new', 'new nguyên': 'new',
                'test': 'test', 'chiết': 'chiet', 'chiet': 'chiet',
                'new mh': 'newmh', 'mat hop': 'newmh', 'mất hộp': 'newmh',
                'mop hop': 'newm', 'móp hộp': 'newm',
                'rach tem': 'newrt', 'rách tem': 'newrt',
                'mop nhe': 'newmn', 'móp nhẹ': 'newmn',
                'xuoc nhe': 'newx', 'xước nhẹ': 'newx',
                'mat hop xuoc': 'newspx', 'mất hộp xước': 'newspx',
            }
            if s in mapping:
                return mapping[s]
            import re as _re
            m = _re.search(r'(\d{2})', s)
            if m and m.group(1) in allowed_status:
                return m.group(1)
            return 'new'

        with transaction.atomic():
            for r_index, row in enumerate(rows[1:], start=2):
                try:
                    def get(col):
                        i = key_indexes.get(col)
                        return row[i] if i is not None and i < len(row) else None

                    name = (get('name') or '').strip() if get('name') else ''
                    if not name:
                        continue

                    brand_name = (get('brand') or '')
                    category_name = (get('category') or '')
                    brand_id_val = get('brand_id')
                    category_id_val = get('category_id')
                    status_val = map_status(get('status'))
                    image_url = (get('image') or '')
                    description = (get('description') or '')
                    ingredients = (get('ingredients') or '')
                    tags_str = (get('tags') or '')

                    import_price = to_int(get('import_price'), 0)
                    original_price = to_int(get('original_price'), 0)
                    discounted_price = get('discounted_price')
                    discount_rate = get('discount_rate')
                    stock_quantity = to_int(get('stock_quantity'), 0)
                    sold_quantity = to_int(get('sold_quantity'), 0)
                    rating = to_float(get('rating'), 4.2)

                    # Chuẩn hóa giá trị
                    if discounted_price in (None, '') and discount_rate not in (None, ''):
                        try:
                            dr = float(discount_rate)
                            discounted_price = int(round(original_price * (100 - dr) / 100))
                        except Exception:
                            discounted_price = original_price
                    else:
                        discounted_price = to_int(discounted_price, original_price)

                    # Tính discount_rate nếu thiếu
                    if discount_rate in (None, ''):
                        try:
                            if original_price and discounted_price <= original_price:
                                discount_rate = round((original_price - discounted_price) * 100.0 / max(original_price, 1), 2)
                            else:
                                discount_rate = 0.0
                        except Exception:
                            discount_rate = 0.0
                    else:
                        discount_rate = float(discount_rate)

                    # Brand: BẮT BUỘC phải có brand_id hợp lệ, không fallback theo tên
                    brand_obj = None
                    try:
                        bid = int(str(brand_id_val)) if brand_id_val not in (None, '') else None
                    except Exception:
                        bid = None
                    if not bid:
                        errors.append({'row': r_index, 'error': 'Thiếu hoặc không hợp lệ brand_id'})
                        continue
                    brand_obj = Brand.objects.filter(id=bid).first()
                    if not brand_obj:
                        errors.append({'row': r_index, 'error': f'Không tìm thấy Brand với id={bid}'})
                        continue

                    # Category: BẮT BUỘC phải có category_id hợp lệ, không fallback theo tên
                    category_obj = None
                    try:
                        cid = int(str(category_id_val)) if category_id_val not in (None, '') else None
                    except Exception:
                        cid = None
                    if not cid:
                        errors.append({'row': r_index, 'error': 'Thiếu hoặc không hợp lệ category_id'})
                        continue
                    category_obj = Category.objects.filter(id=cid).first()
                    if not category_obj:
                        errors.append({'row': r_index, 'error': f'Không tìm thấy Category với id={cid}'})
                        continue

                    # Luôn tạo sản phẩm mới, không kiểm tra trùng
                    product_obj = Product(name=name)
                    is_new = True

                    # Gán thuộc tính
                    product_obj.brand = brand_obj
                    product_obj.category = category_obj
                    product_obj.status = status_val
                    product_obj.image = str(image_url) if image_url else None
                    product_obj.description = str(description)
                    product_obj.ingredients = str(ingredients) if ingredients else ''
                    product_obj.stock_quantity = stock_quantity
                    product_obj.sold_quantity = sold_quantity
                    product_obj.rating = rating
                    product_obj.import_price = import_price
                    product_obj.original_price = original_price
                    product_obj.discounted_price = discounted_price
                    product_obj.discount_rate = discount_rate

                    product_obj.save()

                    # Tags (tên, phân tách bằng , hoặc ;)
                    if tags_str:
                        tag_names = [t.strip() for t in re.split(r'[;,]', str(tags_str)) if t and str(t).strip()]
                        tag_objs = []
                        for tname in tag_names:
                            code = slugify_code(tname)
                            tag, _ = Tag.objects.get_or_create(code=code, defaults={'name': tname, 'description': '', 'status': 'active'})
                            # Đảm bảo tên cập nhật nếu khác
                            if tag.name != tname:
                                tag.name = tname
                                tag.save(update_fields=['name'])
                            tag_objs.append(tag)
                        product_obj.tags.set(tag_objs)

                    created += 1

                except Exception as e:
                    errors.append({'row': r_index, 'error': str(e)})

        return Response({
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total_rows': max(0, len(rows) - 1)
        })


class LatestProductsViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    """
    API endpoint to list the 10 latest products.
    """
    queryset = Product.objects.filter(is_visible=True).select_related('brand', 'category').prefetch_related('tags', 'gifts').order_by('-id')[:10]
    serializer_class = ProductSerializer


# --- Order-related ViewSets ---
class VoucherViewSet(viewsets.ModelViewSet):
    queryset = Voucher.objects.all()
    serializer_class = VoucherSerializer


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().select_related('customer', 'voucher').prefetch_related('items__product')
    serializer_class = OrderSerializer

    def get_permissions(self):
        """
        Set permissions based on action.
        - allow POST for unauthenticated users (for non-logged-in orders)
        - allow GET for unauthenticated users (for admin access)
        - allow PUT, PATCH, DELETE for unauthenticated users (for admin updates)
        """
        # Cho phép tất cả actions cho admin Flask app
        self.permission_classes = [AllowAny]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        """
        Tạo đơn hàng mới với kiểm tra trùng lặp và đảm bảo trạng thái mặc định
        """
        from django.utils import timezone
        from datetime import timedelta
        from django.db import transaction

        # Kiểm tra trùng lặp dựa trên thông tin khách hàng và thời gian
        customer_name = request.data.get('customer_name')
        phone_number = request.data.get('phone_number')

        if customer_name and phone_number:
            # Kiểm tra xem có đơn hàng nào được tạo trong vòng 10 giây với cùng thông tin không
            recent_orders = Order.objects.filter(
                customer_name=customer_name,
                phone_number=phone_number,
                order_date__gte=timezone.now() - timedelta(seconds=10)
            )

            if recent_orders.exists():
                return Response(
                    {'detail': 'Đơn hàng đã được tạo gần đây. Vui lòng đợi một chút.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Đảm bảo trạng thái mặc định
        request.data['status'] = 'pending'
        request.data['is_confirmed'] = False

        # Sử dụng transaction để đảm bảo atomicity
        with transaction.atomic():
            return super().create(request, *args, **kwargs)

    def get_queryset(self):
        """
        Chỉ cho phép người dùng xem đơn hàng của chính họ.
        Cho phép admin Flask app xem tất cả đơn hàng.
        """
        user = self.request.user
        if user.is_authenticated:
            if user.is_staff:
                # Admin Django có thể xem tất cả đơn hàng
                return Order.objects.all().select_related('customer', 'voucher').prefetch_related('items__product')
            else:
                # User thường chỉ xem đơn hàng của chính họ
                return Order.objects.filter(customer=user).select_related('customer', 'voucher').prefetch_related('items__product')
        else:
            # Cho phép Flask admin app truy cập tất cả đơn hàng khi không có authentication
            return Order.objects.all().select_related('customer', 'voucher').prefetch_related('items__product')


class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemReadSerializer


class AnalyticsViewSet(viewsets.GenericViewSet):
    queryset = AnalyticsSnapshot.objects.all()
    serializer_class = AnalyticsSnapshotSerializer
    permission_classes = [AllowAny]

    def _ensure_snapshot_qty_fields(self, snapshots):
        """Recompute snapshots missing newly added fields (qty and phones)."""
        from datetime import date, timedelta
        updated = False
        for snap in snapshots:
            data = snap.data or {}
            has_qty = 'qty_by_category' in data and 'qty_by_brand' in data
            has_phone = 'top_customers_phone' in data and 'top_customer_phone' in data
            if has_qty and has_phone:
                continue
            # Determine start/end from period_key
            period = snap.period
            if period == 'day':
                try:
                    d = date.fromisoformat(snap.period_key)
                except Exception:
                    continue
                payload = { 'period': 'day', 'start': d, 'end': d }
            elif period == 'month':
                try:
                    y, m = map(int, snap.period_key.split('-'))
                    start = date(y, m, 1)
                    if m == 12:
                        end = date(y + 1, 1, 1) - timedelta(days=1)
                    else:
                        end = date(y, m + 1, 1) - timedelta(days=1)
                except Exception:
                    continue
                payload = { 'period': 'month', 'start': start, 'end': end }
            else:
                # Only handle day/month for now
                continue
            serializer = ComputeAnalyticsSerializer(data=payload)
            if serializer.is_valid():
                serializer.save()
                updated = True
        return updated

    def aggregate_snapshots(self, snapshots, period_key):
        """Helper function to aggregate multiple snapshots"""
        if not snapshots:
            return None

        total_revenue = sum(float(snap.total_revenue or 0) for snap in snapshots)
        total_profit = sum(float(snap.total_profit or 0) for snap in snapshots)

        # Aggregate other metrics
        aggregated_data = {
            'labels': [snap.period_key for snap in snapshots],
            'series': [float(snap.total_revenue or 0) for snap in snapshots],
            'total_orders': sum(snap.data.get('total_orders', 0) for snap in snapshots),
            'average_order_value': 0,
            'top_selling': {},
            'top_revenue': {},
            'top_customers': {},
            'top_profit_products': {},
            'revenue_by_region': {},
            'revenue_by_category': {},
            'revenue_by_brand': {},
            'import_total': sum(snap.data.get('import_total', 0) for snap in snapshots),
            'shipping_total': sum(snap.data.get('shipping_total', 0) for snap in snapshots),
        }

        # Calculate average order value
        if aggregated_data['total_orders'] > 0:
            aggregated_data['average_order_value'] = total_revenue / aggregated_data['total_orders']

        # Aggregate top lists
        for snap in snapshots:
            data = snap.data or {}

            # Top selling
            for item in data.get('top_selling', []):
                name = item['name']
                qty = item['qty']
                aggregated_data['top_selling'][name] = aggregated_data['top_selling'].get(name, 0) + qty

            # Top revenue
            for item in data.get('top_revenue', []):
                name = item['name']
                revenue = item['revenue']
                aggregated_data['top_revenue'][name] = aggregated_data['top_revenue'].get(name, 0) + revenue

            # Top customers
            for item in data.get('top_customers', []):
                name = item['name']
                revenue = item['revenue']
                aggregated_data['top_customers'][name] = aggregated_data['top_customers'].get(name, 0) + revenue

            # Top profit products
            for item in data.get('top_profit_products', []):
                name = item['name']
                profit = item['profit']
                aggregated_data['top_profit_products'][name] = aggregated_data['top_profit_products'].get(name, 0) + profit

            # Revenue by region
            for item in data.get('revenue_by_region', []):
                region = item['region']
                revenue = item['revenue']
                aggregated_data['revenue_by_region'][region] = aggregated_data['revenue_by_region'].get(region, 0) + revenue

            # Revenue by category
            for item in data.get('revenue_by_category', []):
                category = item['category']
                revenue = item['revenue']
                aggregated_data['revenue_by_category'][category] = aggregated_data['revenue_by_category'].get(category, 0) + revenue

            # Revenue by brand
            for item in data.get('revenue_by_brand', []):
                brand = item['brand']
                revenue = item['revenue']
                aggregated_data['revenue_by_brand'][brand] = aggregated_data['revenue_by_brand'].get(brand, 0) + revenue

        # Sort and limit top lists
        aggregated_data['top_selling'] = sorted(
            [{'name': name, 'qty': qty} for name, qty in aggregated_data['top_selling'].items()],
            key=lambda x: x['qty'], reverse=True
        )[:10]

        aggregated_data['top_revenue'] = sorted(
            [{'name': name, 'revenue': revenue} for name, revenue in aggregated_data['top_revenue'].items()],
            key=lambda x: x['revenue'], reverse=True
        )[:10]

        aggregated_data['top_customers'] = sorted(
            [{'name': name, 'revenue': revenue} for name, revenue in aggregated_data['top_customers'].items()],
            key=lambda x: x['revenue'], reverse=True
        )[:10]

        aggregated_data['top_profit_products'] = sorted(
            [{'name': name, 'profit': profit} for name, profit in aggregated_data['top_profit_products'].items()],
            key=lambda x: x['profit'], reverse=True
        )[:10]

        aggregated_data['revenue_by_region'] = sorted(
            [{'region': region, 'revenue': revenue} for region, revenue in aggregated_data['revenue_by_region'].items()],
            key=lambda x: x['revenue'], reverse=True
        )[:10]

        aggregated_data['revenue_by_category'] = sorted(
            [{'category': category, 'revenue': revenue} for category, revenue in aggregated_data['revenue_by_category'].items()],
            key=lambda x: x['revenue'], reverse=True
        )[:10]

        aggregated_data['revenue_by_brand'] = sorted(
            [{'brand': brand, 'revenue': revenue} for brand, revenue in aggregated_data['revenue_by_brand'].items()],
            key=lambda x: x['revenue'], reverse=True
        )[:10]

        # Find top customer phone
        top_customer_phone = max(aggregated_data['top_customers'], key=lambda x: x['revenue'])['name'] if aggregated_data['top_customers'] else '-'
        aggregated_data['top_customer_phone'] = top_customer_phone

        return {
            'total_revenue': total_revenue,
            'total_profit': total_profit,
            'data': aggregated_data
        }

    def list(self, request):
        period = request.query_params.get('period', 'month')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Get all unique periods from orders
        from .models import Order
        from datetime import date, timedelta
        from django.db.models import Min, Max

        # Get all confirmed orders
        orders = Order.objects.filter(is_confirmed=True).exclude(status='cancelled')

        # Apply custom date range if provided
        if start_date and end_date:
            try:
                start_date_obj = date.fromisoformat(start_date)
                end_date_obj = date.fromisoformat(end_date)
                orders = orders.filter(order_date__date__gte=start_date_obj, order_date__date__lte=end_date_obj)
                print(f"🔍 Custom date range: {start_date_obj} to {end_date_obj}")
                print(f"📊 Orders found: {orders.count()}")
            except ValueError:
                return Response({'error': 'Invalid date format'}, status=400)

        if not orders.exists():
            print(f"❌ No orders found in the specified range")
            return Response([])

        # For custom date range, return individual day or month snapshots
        if start_date and end_date:
            days_diff = (end_date_obj - start_date_obj).days
            print(f"🔍 Custom range: {start_date_obj} to {end_date_obj} ({days_diff} days)")

            if days_diff > 90:
                # For ranges > 90 days, use month snapshots
                print("📅 Range > 90 days, using month snapshots")

                # Get existing month snapshots in the date range
                month_snapshots = AnalyticsSnapshot.objects.filter(
                    period='month',
                    period_key__gte=start_date_obj.strftime('%Y-%m'),
                    period_key__lte=end_date_obj.strftime('%Y-%m')
                ).order_by('period_key')

                print(f"📊 Found {month_snapshots.count()} month snapshots")

                if month_snapshots.exists():
                    self._ensure_snapshot_qty_fields(month_snapshots)
                    data = AnalyticsSnapshotSerializer(month_snapshots, many=True).data
                    print(f"📊 Returning {len(data)} month snapshots for custom range")
                    return Response(data)
                else:
                    # Create month snapshots for the range
                    current_date = start_date_obj.replace(day=1)
                    end_month = end_date_obj.replace(day=1)

                    while current_date <= end_month:
                        periods_to_compute.append({
                            'period': 'month',
                            'start': current_date,
                            'end': current_date.replace(month=current_date.month + 1) - timedelta(days=1) if current_date.month < 12 else current_date.replace(year=current_date.year + 1, month=1) - timedelta(days=1),
                            'key': current_date.strftime('%Y-%m')
                        })
                        if current_date.month == 12:
                            current_date = current_date.replace(year=current_date.year + 1, month=1)
                        else:
                            current_date = current_date.replace(month=current_date.month + 1)

                    print(f"📅 Created {len(periods_to_compute)} month periods for custom range")
            else:
                # For ranges <= 90 days, use day snapshots
                print("📅 Range <= 90 days, using day snapshots")

                # Get existing day snapshots in the date range
                day_snapshots = AnalyticsSnapshot.objects.filter(
                    period='day',
                    period_key__gte=start_date_obj.strftime('%Y-%m-%d'),
                    period_key__lte=end_date_obj.strftime('%Y-%m-%d')
                ).order_by('period_key')

                print(f"📊 Found {day_snapshots.count()} day snapshots")

                if day_snapshots.exists():
                    self._ensure_snapshot_qty_fields(day_snapshots)
                    data = AnalyticsSnapshotSerializer(day_snapshots, many=True).data
                    print(f"📊 Returning {len(data)} day snapshots for custom range")
                    return Response(data)
                else:
                    # Create day snapshots for the range
                    current_date = start_date_obj
                    while current_date <= end_date_obj:
                        periods_to_compute.append({
                            'period': 'day',
                            'start': current_date,
                            'end': current_date,
                            'key': current_date.strftime('%Y-%m-%d')
                        })
                        current_date += timedelta(days=1)

                    print(f"📅 Created {len(periods_to_compute)} day periods for custom range")
        else:
            # Get date range from orders
            date_range = orders.aggregate(
                min_date=Min('order_date__date'),
                max_date=Max('order_date__date')
            )

            min_date = date_range['min_date']
            max_date = date_range['max_date']

            if not min_date or not max_date:
                return Response([])

            # Generate periods based on actual order dates, not date range
            periods_to_compute = []

            # Get unique dates from orders
            order_dates = orders.values_list('order_date__date', flat=True).distinct().order_by('order_date__date')

            if period == 'day':
                # Always recompute day snapshots based on actual distinct order dates (prevents stale data)
                # Create one period for each day that has orders
                for order_date in order_dates:
                    periods_to_compute.append({
                        'period': 'day',
                        'start': order_date,
                        'end': order_date,
                        'key': order_date.strftime('%Y-%m-%d')
                    })

            elif period == 'week':
                # Group orders by week
                week_groups = {}
                for order_date in order_dates:
                    week_start = order_date - timedelta(days=order_date.weekday())
                    week_key = f"{week_start.year}-W{week_start.strftime('%W')}"
                    if week_key not in week_groups:
                        week_groups[week_key] = {
                            'start': week_start,
                            'end': week_start + timedelta(days=6)
                        }

                for week_key, week_data in week_groups.items():
                    periods_to_compute.append({
                        'period': 'week',
                        'start': week_data['start'],
                        'end': week_data['end'],
                        'key': week_key
                    })

            elif period == 'month':
                # Group orders by month
                month_groups = {}
                for order_date in order_dates:
                    month_key = order_date.strftime('%Y-%m')
                    if month_key not in month_groups:
                        month_start = order_date.replace(day=1)
                        if order_date.month == 12:
                            month_end = order_date.replace(year=order_date.year + 1, month=1, day=1) - timedelta(days=1)
                        else:
                            month_end = order_date.replace(month=order_date.month + 1, day=1) - timedelta(days=1)
                        month_groups[month_key] = {
                            'start': month_start,
                            'end': month_end
                        }

                for month_key, month_data in month_groups.items():
                    periods_to_compute.append({
                        'period': 'month',
                        'start': month_data['start'],
                        'end': month_data['end'],
                        'key': month_key
                    })

            elif period == 'year':
                # Group orders by year
                year_groups = {}
                for order_date in order_dates:
                    year_key = str(order_date.year)
                    if year_key not in year_groups:
                        year_groups[year_key] = {
                            'start': date(order_date.year, 1, 1),
                            'end': date(order_date.year, 12, 31)
                        }

                for year_key, year_data in year_groups.items():
                    periods_to_compute.append({
                        'period': 'year',
                        'start': year_data['start'],
                        'end': year_data['end'],
                        'key': year_key
                    })

            # For regular periods, delete existing snapshots
            period_to_delete = period

        # Delete all existing snapshots for this period type to avoid confusion
        AnalyticsSnapshot.objects.filter(period=period_to_delete).delete()

        # Compute snapshots for all periods
        snapshots = []
        print(f"🔄 Computing {len(periods_to_compute)} periods...")
        for period_data in periods_to_compute:
            print(f"📅 Processing period: {period_data}")
            # Create new snapshot
            serializer = ComputeAnalyticsSerializer(data=period_data)
            if serializer.is_valid():
                snapshot = serializer.save()
                snapshots.append(snapshot)
                print(f"✅ Created snapshot: {snapshot.period_key} - Revenue: {snapshot.total_revenue}")
            else:
                print(f"❌ Serializer errors: {serializer.errors}")

        # Return snapshots sorted by period_key descending (newest first)
        snapshots.sort(key=lambda x: x.period_key, reverse=True)
        data = AnalyticsSnapshotSerializer(snapshots, many=True).data

        # Debug: Log the snapshots being returned
        print(f"📊 Returning {len(data)} snapshots for period {period}")
        for snap in data:
            print(f"  - {snap['period']} {snap['period_key']}: {snap['total_revenue']}đ")

        return Response(data)

    @action(detail=False, methods=['post'], url_path='compute')
    def compute(self, request):
        serializer = ComputeAnalyticsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        snapshot = serializer.save()
        return Response(AnalyticsSnapshotSerializer(snapshot).data, status=status.HTTP_201_CREATED)

class OrderByCodeView(APIView):
    """
    API để tra cứu đơn hàng bằng mã đơn hàng.
    """
    permission_classes = [AllowAny] # Cho phép bất kỳ ai cũng có thể truy cập

    def get(self, request, order_code, format=None):
        """
        Lấy thông tin đơn hàng dựa trên mã đơn hàng.
        """
        try:
            order = Order.objects.get(order_code=order_code)
            serializer = OrderSerializer(order)
            return Response(serializer.data)
        except Order.DoesNotExist:
            return Response(
                {"detail": "Mã đơn hàng không tồn tại hoặc không hợp lệ."},
                status=status.HTTP_404_NOT_FOUND
            )


# --- View Đăng Ký và Đăng Nhập Xã Hội ---
class RegistrationView(generics.CreateAPIView):
    """
    Registration endpoint to create a new user.
    """
    serializer_class = RegistrationSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            'message': 'Đăng ký thành công',
            'access_token': str(refresh.access_token),
            'refresh_token': str(refresh),
            'user': UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)


class GoogleSocialAuthView(GenericAPIView):
    serializer_class = GoogleSocialAuthSerializer
    permission_classes = (AllowAny,)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Corrected line: Get the user object directly from the validated data
        # of the 'auth_token' field.
        user = serializer.validated_data['auth_token']

        # Now 'user' is the correct User model instance.
        refresh = RefreshToken.for_user(user)
        response_data = {
            'user': UserSerializer(user).data,
            'access_token': str(refresh.access_token),
            'refresh_token': str(refresh),
        }
        return Response(response_data, status=status.HTTP_200_OK)


# class PhoneNumberLoginView(APIView):
#     """
#     API View to handle login requests with phone number and password.
#     Returns user info and JWT tokens upon successful login.
#     """
#     permission_classes = [AllowAny]

#     def post(self, request, *args, **kwargs):
#         serializer = PhoneNumberLoginSerializer(data=request.data, context={'request': request})
#         if serializer.is_valid():
#             # Thay đổi dòng này từ serializer.validated_data sang serializer.instance
#             return Response(serializer.to_representation(serializer.instance), status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = LoginSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            # Sửa lỗi ở đây
            return Response(serializer.to_representation(serializer.instance), status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CartViewSet(viewsets.GenericViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Cart.objects.filter(user=self.request.user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset().prefetch_related('items__product')
        cart_instance = queryset.first()
        if not cart_instance:
            cart_instance = Cart.objects.create(user=request.user)
            serializer = self.get_serializer(cart_instance)
        else:
            serializer = self.get_serializer(cart_instance)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='add-item', serializer_class=CartItemSerializer)
    def add_item_to_cart(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        product_id = serializer.validated_data['product'].id
        quantity = serializer.validated_data['quantity']

        cart, created = Cart.objects.get_or_create(user=self.request.user)

        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            product_id=product_id,
            defaults={'quantity': quantity}
        )
        if not created:
            cart_item.quantity += quantity
            cart_item.save()

        return Response(CartItemSerializer(cart_item).data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['put'], url_path='update-item', serializer_class=CartItemSerializer)
    def update_cart_item_quantity(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        cart = Cart.objects.get(user=self.request.user)
        product_id = serializer.validated_data['product'].id
        quantity = serializer.validated_data['quantity']

        try:
            cart_item = CartItem.objects.get(cart=cart, product_id=product_id)
            cart_item.quantity = quantity
            cart_item.save()
            return Response(CartItemSerializer(cart_item).data, status=status.HTTP_200_OK)
        except CartItem.DoesNotExist:
            return Response({"detail": "Sản phẩm không có trong giỏ hàng."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['delete'], url_path='remove-item')
    def remove_cart_item(self, request):
        product_id = request.data.get('product_id')

        try:
            cart = Cart.objects.get(user=self.request.user)
            cart_item = CartItem.objects.get(cart=cart, product_id=product_id)
            cart_item.delete()
        except (Cart.DoesNotExist, CartItem.DoesNotExist):
            return Response({"detail": "Giỏ hàng hoặc sản phẩm không được tìm thấy."}, status=status.HTTP_404_NOT_FOUND)

        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='item-count')
    def get_cart_item_count(self, request):
        try:
            cart = Cart.objects.get(user=self.request.user)
            total_items = CartItem.objects.filter(cart=cart).count()
            return Response({'item_count': total_items}, status=status.HTTP_200_OK)
        except Cart.DoesNotExist:
            return Response({'item_count': 0}, status=status.HTTP_200_OK)


class UserProfileView(generics.RetrieveUpdateAPIView):
    """
    View để lấy và cập nhật hồ sơ của người dùng đã đăng nhập.
    """
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

        return Response({
            "message": "Thông tin cá nhân đã được cập nhật thành công.",
            "user": serializer.data
        })




# --- Admin ViewSet ---
class AdminViewSet(viewsets.ModelViewSet):
    queryset = Admin.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['username', 'email', 'phone_number']





# --- Lucky Event Views ---
class LuckyEventViewSet(viewsets.ModelViewSet):
    queryset = LuckyEvent.objects.all().order_by('-id')
    serializer_class = LuckyEventSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='active')
    def get_active_event(self, request):
        from django.utils import timezone
        now = timezone.now()
        event = LuckyEvent.objects.filter(is_active=True, start_at__lte=now, end_at__gte=now).order_by('-id').first()
        if not event:
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response(LuckyEventSerializer(event).data)

    @action(detail=True, methods=['get'], url_path='winners')
    def winners(self, request, pk=None):
        event = self.get_object()
        winners_qs = LuckyWinner.objects.filter(event=event).select_related('participant', 'prize')
        return Response(LuckyWinnerSerializer(winners_qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='finalize')
    def finalize(self, request, pk=None):
        """Chốt kết quả: lấy 2 số cuối từ xổ số hoặc nhận lucky_number thủ công."""
        from django.utils import timezone
        event = self.get_object()
        # lấy lucky_number từ request nếu có (fallback thủ công)
        manual = (request.data or {}).get('lucky_number')
        if manual and isinstance(manual, str) and len(manual) == 2 and manual.isdigit():
            event.lucky_number = manual
            event.save(update_fields=['lucky_number'])

        if not event.lucky_number:
            return Response({'detail': 'Chưa có lucky_number'}, status=400)

        # tìm tất cả người thắng theo số may mắn, sắp xếp theo thời gian gửi
        participants = list(LuckyParticipant.objects.filter(event=event, chosen_number=event.lucky_number).order_by('submitted_at'))
        if not participants:
            return Response({'detail': 'Không có người thắng cho số này'}, status=200)

        prizes = list(event.prizes.order_by('order', 'id'))
        if not prizes:
            return Response({'detail': 'Chưa cấu hình giải thưởng cho sự kiện'}, status=400)

        # Gán lần lượt từng prize cho người thắng theo thứ tự gửi
        LuckyWinner.objects.filter(event=event).delete()
        created = []
        for idx, prize in enumerate(prizes):
            if idx >= len(participants):
                break
            winner = LuckyWinner.objects.create(event=event, participant=participants[idx], prize=prize)
            created.append(winner)

        return Response({'detail': 'Đã chốt kết quả', 'winners': LuckyWinnerSerializer(created, many=True).data})


class LuckyParticipantViewSet(viewsets.ModelViewSet):
    queryset = LuckyParticipant.objects.all().select_related('event')
    permission_classes = [AllowAny]

    def get_serializer_class(self):
        if self.action in ['create']:
            return LuckyParticipantCreateSerializer
        return LuckyParticipantSerializer

class BlogAPIView(viewsets.ModelViewSet):
    """
    API View để lấy danh sách blog, chi tiết, cập nhật, hoặc xóa một bài blog.
    Có thể lọc danh sách blog theo tag.
    """
    serializer_class = BlogSerializer
    pagination_class = NewsPagination
    lookup_field = 'pk'  # Sử dụng primary key để tìm kiếm blog

    def get_queryset(self):
        """
        Tùy chỉnh queryset để có thể lọc blog theo tag.
        Nếu có tham số 'tag' trong request, sẽ lọc các blog có tag tương ứng.
        """
        queryset = Blog.objects.filter(is_active=True)
        tag = self.request.query_params.get('tag', None)
        if tag is not None:
            # Lọc các bài blog có tag trùng khớp
            queryset = queryset.filter(tag=tag.upper())
        return queryset

    def retrieve(self, request, *args, **kwargs):
        """
        Lấy chi tiết một bài blog và tăng lượt xem khi được xem.
        """
        try:
            instance = self.get_object()
            # Tăng lượt xem của bài blog lên 1 mỗi khi được truy cập
            instance.views += 1
            instance.save()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Blog.DoesNotExist:
            return Response({"detail": "Bài viết không tồn tại."}, status=status.HTTP_404_NOT_FOUND)

    def update(self, request, *args, **kwargs):
        """
        Cập nhật toàn bộ tài nguyên Blog (sử dụng PUT)
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance = self.get_object()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        """
        Cập nhật một phần tài nguyên Blog (sử dụng PATCH)
        """
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """
        Xóa một bài Blog
        """
        try:
            instance = self.get_object()
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Blog.DoesNotExist:
            return Response({"detail": "Bài viết không tồn tại."}, status=status.HTTP_404_NOT_FOUND)

class LuckyWinnerViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet cho LuckyWinner - chỉ đọc
    """
    queryset = LuckyWinner.objects.all()
    serializer_class = LuckyWinnerSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        """
        Lọc theo event nếu có
        """
        queryset = LuckyWinner.objects.select_related('participant', 'prize', 'event')
        event_id = self.request.query_params.get('event', None)
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        return queryset


class CustomerLeadViewSet(viewsets.ModelViewSet):
    queryset = CustomerLead.objects.all().order_by('-updated_at')
    serializer_class = CustomerLeadSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        """Prevent duplicate phone numbers by using CustomerLead.upsert.
        If no phone is provided, skip (do not create).
        """
        name = (request.data.get('name') or '').strip()
        phone = (request.data.get('phone') or '').strip()
        email = (request.data.get('email') or '').strip()
        address = (request.data.get('address') or '').strip()

        lead = CustomerLead.upsert(name=name, phone=phone, email=email, address=address)
        if not lead:
            return Response({'detail': 'Thiếu SĐT'}, status=status.HTTP_400_BAD_REQUEST)
        ser = self.get_serializer(lead)
        # 201 if newly created else 200; khó phân biệt an toàn => trả 200 để tránh lỗi phía client
        return Response(ser.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='sync')
    def sync(self, request):
        created_or_updated = 0
        try:
            for o in Order.objects.all().only('customer_name','phone_number','email','street','ward','district','province'):
                addr = ', '.join(filter(None, [o.street, o.ward, o.district, o.province]))
                if CustomerLead.upsert(o.customer_name, o.phone_number, o.email, addr):
                    created_or_updated += 1
            for p in LuckyParticipant.objects.all().only('name','zalo_phone','email','address'):
                if CustomerLead.upsert(p.name, p.zalo_phone, p.email, p.address):
                    created_or_updated += 1
            for u in Customer.objects.all().only('name','phone_number','email','address'):
                if CustomerLead.upsert(u.name, u.phone_number, u.email, u.address):
                    created_or_updated += 1
        except Exception as e:
            return Response({'detail': f'Lỗi đồng bộ: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'detail': 'Đã đồng bộ', 'count': created_or_updated})

    @action(detail=False, methods=['post'], url_path='dedupe')
    def dedupe(self, request):
        """Xóa các bản ghi trùng SĐT, giữ lại bản ghi cập nhật gần nhất."""
        from collections import defaultdict
        buckets = defaultdict(list)
        for lead in CustomerLead.objects.exclude(phone__isnull=True).exclude(phone=''):
            buckets[lead.phone.strip()].append(lead)
        removed = 0
        for phone, items in buckets.items():
            if len(items) <= 1:
                continue
            # giữ lại item mới nhất theo updated_at
            items.sort(key=lambda x: (x.updated_at or x.created_at or timezone.now()), reverse=True)
            keep = items[0]
            for d in items[1:]:
                d.delete(); removed += 1
        return Response({'detail': 'Đã xóa trùng', 'removed': removed})


# --- CTV (Affiliate) Views ---
class CTVApplicationViewSet(viewsets.ModelViewSet):
    queryset = CTVApplication.objects.all().order_by('-created_at')
    serializer_class = CTVApplicationSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='check-code')
    def check_code(self, request):
        code = (request.query_params.get('code') or '').strip()
        if not code:
            return Response({'detail': 'Thiếu mã cần kiểm tra'}, status=400)
        import re
        if len(code) < 3 or len(code) > 20 or not re.match(r'^[A-Za-z0-9_-]+$', code):
            return Response({'available': False, 'reason': 'Mã không hợp lệ (3-20 ký tự, chữ/số/_/-)'}, status=200)
        exists = CTV.objects.filter(code__iexact=code).exists()
        return Response({'available': not exists})

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """Duyệt đơn CTV: tạo CTV + ví, gán level mặc định, set status=approved"""
        app = self.get_object()
        if app.status == 'approved':
            return Response({'detail': 'Đơn đã được duyệt trước đó'}, status=400)

        # Lấy password từ request
        password_text = request.data.get('password_text', '')
        if not password_text:
            return Response({'detail': 'Vui lòng cung cấp mật khẩu cho CTV'}, status=400)

        # Kiểm tra SĐT đã tồn tại chưa
        if CTV.objects.filter(phone=app.phone).exists():
            return Response({'detail': f'Số điện thoại {app.phone} đã được sử dụng'}, status=400)

        # Chọn mã: ưu tiên desired_code, fallback theo phone
        desired = (getattr(app, 'desired_code', '') or '').strip()
        if desired and CTV.objects.filter(code__iexact=desired).exists():
            return Response({'detail': f'Mã CTV "{desired}" đã được sử dụng'}, status=400)

        code = desired or f"CTV{app.phone[-4:] if app.phone else ''}"
        base = code or 'CTV'
        candidate = base
        suffix = 1
        while CTV.objects.filter(code__iexact=candidate).exists():
            candidate = f"{base}{suffix}"
            suffix += 1

        # Lấy level mặc định
        level = CTVLevel.objects.order_by('id').first()

        ctv = CTV.objects.create(
            code=candidate,
            desired_code=desired,  # Lưu mã mong muốn
            full_name=app.full_name,
            phone=app.phone,
            email=app.email,
            address=app.address,
            bank_name=app.bank_name,
            bank_number=app.bank_number,
            bank_holder=app.bank_holder,
            cccd_front_url=app.cccd_front_url,
            cccd_back_url=app.cccd_back_url,
            password_text=password_text,  # Lưu mật khẩu
            level=level,
            is_active=True,
        )
        CTVWallet.objects.get_or_create(ctv=ctv)

        app.status = 'approved'
        app.agreed = True
        app.save(update_fields=['status', 'agreed'])

        return Response({'detail': 'Đã duyệt đơn và tạo tài khoản CTV', 'ctv': CTVSerializer(ctv).data})


class CTVLevelViewSet(viewsets.ModelViewSet):
    queryset = CTVLevel.objects.all()
    serializer_class = CTVLevelSerializer
    permission_classes = [AllowAny]

    def update(self, request, *args, **kwargs):
        """Override update method to ensure it works properly"""
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        """Override partial_update method to ensure it works properly"""
        return super().partial_update(request, *args, **kwargs)


class CTVViewSet(viewsets.ModelViewSet):
    queryset = CTV.objects.all().select_related('level').prefetch_related('wallet')
    serializer_class = CTVSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='by-code')
    def by_code(self, request):
        code = (request.query_params.get('code') or '').strip()
        if not code:
            return Response({'detail': 'Thiếu mã CTV'}, status=400)
        ctv = CTV.objects.filter(code__iexact=code).select_related('level').first()
        if not ctv:
            return Response({'detail': 'Không tìm thấy CTV với mã này'}, status=404)
        return Response(CTVSerializer(ctv).data)

    @action(detail=False, methods=['get'], url_path='by-phone')
    def by_phone(self, request):
        phone = (request.query_params.get('phone') or '').strip()
        email = (request.query_params.get('email') or '').strip()
        if not phone and not email:
            return Response({'detail': 'Thiếu phone hoặc email'}, status=400)
        q = CTV.objects.all()
        if phone:
            q = q.filter(phone=phone)
        if email:
            q = q.filter(email=email)
        ctv = q.select_related('level').first()
        if not ctv:
            return Response({'detail': 'Không tìm thấy CTV'}, status=404)
        return Response(CTVSerializer(ctv).data)

    @action(detail=False, methods=['post'], url_path='login')
    def login(self, request):
        """Đăng nhập CTV bằng phone và password_text"""
        phone = (request.data.get('phone') or '').strip()
        password = (request.data.get('password') or '').strip()

        if not phone or not password:
            return Response({'detail': 'Thiếu phone hoặc password'}, status=400)

        try:
            ctv = CTV.objects.get(phone=phone, is_active=True)
            if not ctv.password_text:
                return Response({'detail': 'CTV chưa có mật khẩu. Liên hệ admin.'}, status=400)

            if ctv.password_text != password:
                return Response({'detail': 'Mật khẩu không đúng'}, status=400)

            return Response({
                'detail': 'Đăng nhập thành công',
                'ctv': CTVSerializer(ctv).data
            })
        except CTV.DoesNotExist:
            return Response({'detail': 'Không tìm thấy CTV với số điện thoại này'}, status=404)

    @action(detail=True, methods=['get'], url_path='wallet')
    def wallet(self, request, pk=None):
        ctv = self.get_object()
        wallet, _ = CTVWallet.objects.get_or_create(ctv=ctv)
        return Response(CTVWalletSerializer(wallet).data)

    @action(detail=True, methods=['post'], url_path='withdraw')
    def withdraw(self, request, pk=None):
        ctv = self.get_object()
        amount = request.data.get('amount')
        try:
            amt = float(amount)
        except Exception:
            return Response({'detail': 'Số tiền không hợp lệ'}, status=400)
        if amt <= 0:
            return Response({'detail': 'Số tiền phải > 0'}, status=400)
        wallet, _ = CTVWallet.objects.get_or_create(ctv=ctv)
        if amt > float(wallet.balance):
            return Response({'detail': 'Số dư không đủ'}, status=400)
        # move to pending
        wallet.balance = float(wallet.balance) - amt
        wallet.pending = float(wallet.pending) + amt
        wallet.save()
        CTVWithdrawal.objects.create(ctv=ctv, amount=amt, status='pending')
        return Response({'detail': 'Đã tạo yêu cầu rút, chờ duyệt', 'wallet': CTVWalletSerializer(wallet).data})

    @action(detail=True, methods=['get'], url_path='withdrawals')
    def withdrawals(self, request, pk=None):
        ctv = self.get_object()
        qs = ctv.withdrawals.all().order_by('-requested_at')
        return Response(CTVWithdrawalSerializer(qs, many=True).data)

    @action(detail=True, methods=['get'], url_path='stats')
    def stats(self, request, pk=None):
        """Tổng quan: tổng đơn, doanh thu, hoa hồng đã ghi nhận theo ví."""
        ctv = self.get_object()
        code = ctv.code
        orders = Order.objects.filter(collaborator_code=code, is_confirmed=True).exclude(status='cancelled')
        total_orders = orders.count()
        total_revenue = float(sum([o.total_amount or 0 for o in orders]))
        # Commission = số dư ví + pending + tổng số đã rút
        wallet, _ = CTVWallet.objects.get_or_create(ctv=ctv)
        withdrawn_total = float(sum([w.amount for w in ctv.withdrawals.filter(status='approved')]))
        data = {
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'commission_balance': float(wallet.balance),
            'commission_pending': float(wallet.pending),
            'commission_withdrawn': withdrawn_total,
        }
        return Response(data)

    @action(detail=True, methods=['get'], url_path='commissions')
    def commissions(self, request, pk=None):
        """Danh sách đơn có mã CTV và hoa hồng ước tính theo level hiện tại.
        Lợi nhuận = Tổng tiền - Phí ship - Giảm giá - Giá nhập
        Hoa hồng = Lợi nhuận * phần trăm hoa hồng
        """
        ctv = self.get_object()
        percent = float(ctv.level.commission_percent if ctv.level else 10.0) / 100.0
        orders = Order.objects.filter(collaborator_code=ctv.code, is_confirmed=True).exclude(status='cancelled').prefetch_related('items__product')
        items = []
        total_revenue_sum = 0.0

        for o in orders:
            # Tính tổng giá nhập của đơn hàng
            total_import_cost = 0.0
            order_items_data = []

            for it in o.items.all():
                # Giá nhập (đã tính theo đơn vị nghìn đồng)
                import_price = float(getattr(it.product, 'import_price', 0) or 0)
                quantity = float(it.quantity or 0)
                item_import_cost = import_price * quantity
                total_import_cost += item_import_cost

                order_items_data.append({
                    'product_name': getattr(it.product, 'name', f"#{it.product_id}"),
                    'quantity': quantity,
                    'import_price': import_price,
                    'item_import_cost': item_import_cost
                })

            # Tính lợi nhuận cho toàn bộ đơn hàng
            # Tổng tiền - Phí ship - Giảm giá - Giá nhập
            order_total = float(o.total_amount or 0)
            shipping_fee = float(o.shipping_fee or 0)
            discount_applied = float(o.discount_applied or 0)

            # Tất cả giá trị đều đã ở đơn vị nghìn đồng
            order_profit = order_total - shipping_fee - discount_applied - total_import_cost

            # Tính hoa hồng cho đơn hàng
            order_commission = max(0.0, order_profit * percent)

            # Tính doanh thu bán được
            total_revenue_sum += order_total

            # Thêm vào danh sách items
            items.append({
                'order_code': getattr(o, 'order_code', f"#{o.id}"),
                'product_name': f"{len(order_items_data)} sản phẩm",  # Tổng số sản phẩm trong đơn
                'quantity': 1,  # 1 đơn hàng
                'profit': round(order_profit, 2),
                'commission': round(order_commission, 2),
            })

        # cập nhật tổng doanh thu bán được cho CTV
        try:
            ctv.total_revenue = float(ctv.total_revenue or 0)  # ensure exists
            ctv.total_revenue = float(total_revenue_sum)
            ctv.save(update_fields=['total_revenue'])
        except Exception:
            pass
        return Response(items)


class CTVWithdrawalViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CTVWithdrawal.objects.all().select_related('ctv')
    serializer_class = CTVWithdrawalSerializer
    permission_classes = [AllowAny]

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """Duyệt yêu cầu rút tiền"""
        withdrawal = self.get_object()
        if withdrawal.status != 'pending':
            return Response({'detail': 'Yêu cầu đã được xử lý'}, status=400)

        note = request.data.get('note', '')

        # Cập nhật trạng thái
        withdrawal.status = 'approved'
        withdrawal.processed_at = timezone.now()
        withdrawal.note = note
        withdrawal.save()

        # Cập nhật ví CTV: chuyển từ pending sang withdrawn
        wallet = withdrawal.ctv.wallet
        wallet.pending = float(wallet.pending) - float(withdrawal.amount)
        wallet.save()

        return Response({'detail': 'Đã duyệt yêu cầu rút tiền'})

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Từ chối yêu cầu rút tiền"""
        withdrawal = self.get_object()
        if withdrawal.status != 'pending':
            return Response({'detail': 'Yêu cầu đã được xử lý'}, status=400)

        note = request.data.get('note', '')

        # Cập nhật trạng thái
        withdrawal.status = 'rejected'
        withdrawal.processed_at = timezone.now()
        withdrawal.note = note
        withdrawal.save()

        # Hoàn lại tiền vào ví CTV: chuyển từ pending về balance
        wallet = withdrawal.ctv.wallet
        wallet.pending = float(wallet.pending) - float(withdrawal.amount)
        wallet.balance = float(wallet.balance) + float(withdrawal.amount)
        wallet.save()

        return Response({'detail': 'Đã từ chối yêu cầu rút tiền'})


# --- CTV Pages (Templates) ---
def ctv_login_page(request):
    return render(request, 'ctv_login.html')


def ctv_dashboard_page(request):
    return render(request, 'ctv_dashboard.html')


def ctv_wallet_page(request):
    return render(request, 'ctv_wallet.html')


def ctv_orders_page(request):
    return render(request, 'ctv_orders.html')


def ctv_profile_page(request):
    return render(request, 'ctv_profile.html')


def ctv_place_order_page(request):
    return render(request, 'ctv_place_order.html')


def ctv_resources_page(request):
    return render(request, 'ctv_resources.html')


def admin_resources_page(request):
    return render(request, 'admin_resources.html')


class MarketingResourceViewSet(viewsets.ModelViewSet):
    queryset = MarketingResource.objects.all().order_by('-created_at')
    serializer_class = MarketingResourceSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['resource_type', 'is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'download_count']
    ordering = ['-created_at']
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        """Override để admin có thể xem tất cả, CTV chỉ xem active"""
        if self.request.path.startswith('/admin/'):
            return MarketingResource.objects.all().order_by('-created_at')
        return MarketingResource.objects.filter(is_active=True).order_by('-created_at')

    def create(self, request, *args, **kwargs):
        """Override create để xử lý upload file lên Cloudinary hoặc tạo record với URL có sẵn"""
        try:
            file = request.FILES.get('file')

            if file:
                # Upload file lên Cloudinary
                import cloudinary
                import cloudinary.uploader
                from datetime import datetime

                # Upload to Cloudinary with image optimization for marketing resources
                upload_result = cloudinary.uploader.upload(
                    file,
                    folder="marketing_resources",
                    public_id=f"resource_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.name.split('.')[0]}",
                    resource_type="image",
                    # Image optimization settings
                    quality="auto:good",  # Chất lượng tốt cho marketing materials
                    fetch_format="auto",  # Tự động chọn format tối ưu
                    width=2000,  # Giới hạn chiều rộng tối đa cho marketing
                    height=2000,  # Giới hạn chiều cao tối đa cho marketing
                    crop="limit",  # Giữ nguyên tỷ lệ
                    flags="progressive",  # Tạo ảnh progressive
                    transformation=[
                        {"width": 2000, "height": 2000, "crop": "limit"},
                        {"quality": "auto:good"},
                        {"fetch_format": "auto"}
                    ]
                )

                # Tạo MarketingResource với URL từ Cloudinary
                resource_data = {
                    'name': request.data.get('name', file.name),
                    'description': request.data.get('description', ''),
                    'resource_type': request.data.get('resource_type', 'new_product'),
                    'file_url': upload_result['secure_url'],
                    'thumbnail_url': upload_result['secure_url'],
                    'is_active': str(request.data.get('is_active', 'true')).lower() == 'true',
                    'file_size': file.size
                }
            else:
                # Tạo record với dữ liệu có sẵn (từ frontend đã upload lên Cloudinary)
                resource_data = {
                    'name': request.data.get('name', ''),
                    'description': request.data.get('description', ''),
                    'resource_type': request.data.get('resource_type', 'new_product'),
                    'file_url': request.data.get('file_url', ''),
                    'thumbnail_url': request.data.get('thumbnail_url', ''),
                    'is_active': str(request.data.get('is_active', 'true')).lower() == 'true',
                    'file_size': request.data.get('file_size', 0)
                }

            serializer = self.get_serializer(data=resource_data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)

            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=201, headers=headers)

        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def destroy(self, request, *args, **kwargs):
        """Override destroy để xóa file trên Cloudinary khi xóa record"""
        instance = self.get_object()

        # Xóa file trên Cloudinary nếu tồn tại
        if instance.file_url and 'cloudinary.com' in instance.file_url:
            try:
                import cloudinary
                import cloudinary.uploader

                # Extract public_id from Cloudinary URL
                # URL format: https://res.cloudinary.com/cloud_name/image/upload/v1234567890/folder/filename.jpg
                url_parts = instance.file_url.split('/')
                if len(url_parts) >= 8:  # Ensure we have enough parts
                    # Get the part after 'upload/v1234567890/'
                    upload_index = url_parts.index('upload')
                    if upload_index + 2 < len(url_parts):
                        public_id = '/'.join(url_parts[upload_index + 2:])  # Get folder/filename
                        public_id = public_id.split('.')[0]  # Remove extension

                        # Delete from Cloudinary
                        result = cloudinary.uploader.destroy(public_id)
                        if result.get('result') == 'ok':
                            print(f"✅ Đã xóa file trên Cloudinary: {public_id}")
                        else:
                            print(f"⚠️ Không thể xóa file trên Cloudinary: {public_id}")

            except Exception as e:
                print(f"❌ Lỗi khi xóa file trên Cloudinary: {e}")

        # Xóa record trong database
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='download')
    def download(self, request, pk=None):
        """Tăng số lượt tải xuống khi CTV tải file"""
        resource = self.get_object()
        resource.download_count += 1
        resource.save(update_fields=['download_count'])

        return Response({
            'message': 'Download count updated',
            'download_count': resource.download_count
        })

    @action(detail=False, methods=['get'], url_path='by-type')
    def by_type(self, request):
        """Lấy tài nguyên theo loại"""
        resource_type = request.query_params.get('type')
        if resource_type:
            resources = self.queryset.filter(resource_type=resource_type)
            serializer = self.get_serializer(resources, many=True)
            return Response(serializer.data)
        return Response([])

    @action(detail=False, methods=['get'], url_path='products')
    def get_product_images(self, request):
        """Lấy ảnh sản phẩm từ products"""
        from .models import Product
        products = Product.objects.filter(status='new', is_visible=True).select_related('brand', 'category')
        product_data = []

        for product in products:
            if product.image:
                product_data.append({
                    'id': f"product_{product.id}",
                    'name': product.name,
                    'description': f"Sản phẩm {product.brand.name if product.brand else 'Không xác định'}",
                    'resource_type': 'product',
                    'file_url': product.image,
                    'thumbnail_url': product.image,
                    'product_id': product.id,
                    'product_name': product.name,
                    'is_image': True,
                    'file_extension': product.image.split('.')[-1].lower() if '.' in product.image else 'jpg'
                })

        # Tắt pagination bằng cách override get_paginated_response
        self.pagination_class = None
        return Response(product_data)


class ProductImagesAPIView(APIView):
    """API View riêng để lấy ảnh sản phẩm không bị pagination"""
    permission_classes = [AllowAny]

    def get(self, request):
        """Lấy tất cả ảnh sản phẩm"""
        from .models import Product
        products = Product.objects.filter(is_visible=True).select_related('brand', 'category')
        print(f"🔍 DEBUG: Total products: {products.count()}")
        new_products = products.filter(status='new')
        print(f"🔍 DEBUG: Products with status='new': {new_products.count()}")

        product_data = []
        products_with_image = 0

        for product in products:
            if product.image and product.status == 'new':
                products_with_image += 1
                product_data.append({
                    'id': f"product_{product.id}",
                    'name': product.name,
                    'description': f"Sản phẩm {product.brand.name if product.brand else 'Không xác định'}",
                    'resource_type': 'product',
                    'file_url': product.image,
                    'thumbnail_url': product.image,
                    'product_id': product.id,
                    'product_name': product.name,
                    'is_image': True,
                    'file_extension': product.image.split('.')[-1].lower() if '.' in product.image else 'jpg'
                })

        print(f"🔍 DEBUG: Products with image: {products_with_image}")
        print(f"🔍 DEBUG: Returning {len(product_data)} products")
        return Response(product_data)


